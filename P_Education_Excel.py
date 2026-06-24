"""PROJECT 3 — EDUCATION  |  STEP 3: Excel Report"""
import pandas as pd,sqlite3
from openpyxl import load_workbook
from openpyxl.styles import PatternFill,Font,Alignment,Border,Side
from openpyxl.utils import get_column_letter

conn=sqlite3.connect('/home/claude/project3_education/education.db')
df=pd.read_sql_query("SELECT * FROM students",conn); conn.close()
df['attendance_pct']=df['attendance_pct'].fillna(df['attendance_pct'].median())
df['study_hours_day']=df['study_hours_day'].fillna(df['study_hours_day'].median())
df['prev_gpa']=df['prev_gpa'].fillna(df['prev_gpa'].median())
df['grade_letter']=pd.cut(df['final_grade'],bins=[0,4,5.5,7,8.5,10],labels=['F','D','C','B','A'])
df['attendance_band']=pd.cut(df['attendance_pct'],bins=[0,59,74,89,100],labels=['<60%','60-74%','75-89%','90%+'])

kpi=pd.DataFrame({'Metric':['Total Students','Dropout Risk Rate (%)','Avg Final Grade (/10)',
    'Avg Attendance (%)','Avg Study Hours/Day','At-Risk Students (attendance<60%)'],
    'Value':[len(df),round(df['dropout_risk'].mean()*100,1),round(df['final_grade'].mean(),2),
             round(df['attendance_pct'].mean(),1),round(df['study_hours_day'].mean(),1),
             int((df['attendance_pct']<60).sum())]})
dept=df.groupby('department').agg(Students=('student_id','count'),Avg_Grade=('final_grade','mean'),
    Avg_Attendance=('attendance_pct','mean'),Dropout_Rate=('dropout_risk','mean'),
    Avg_Study_Hours=('study_hours_day','mean')).round(2).reset_index()
dept['Dropout_Rate']=(dept['Dropout_Rate']*100).round(1)
loc=df.groupby('location').agg(Students=('student_id','count'),Avg_Grade=('final_grade','mean'),
    Dropout_Rate=('dropout_risk','mean'),Avg_Attendance=('attendance_pct','mean')).round(2).reset_index()
loc['Dropout_Rate']=(loc['Dropout_Rate']*100).round(1)
par=df.groupby('parental_edu').agg(Students=('student_id','count'),Avg_Grade=('final_grade','mean'),
    Dropout_Rate=('dropout_risk','mean')).round(2).reset_index()
par['Dropout_Rate']=(par['Dropout_Rate']*100).round(1)

path='/home/claude/project3_education/Education_Report.xlsx'
with pd.ExcelWriter(path,engine='openpyxl') as w:
    kpi.to_excel(w,sheet_name='KPI Summary',index=False)
    dept.to_excel(w,sheet_name='Dept Performance',index=False)
    loc.to_excel(w,sheet_name='Location Analysis',index=False)
    par.to_excel(w,sheet_name='Parental Edu Impact',index=False)

wb=load_workbook(path)
WHITE='FFFFFF';LGRAY='E0EEF5'
def style_sheet(ws,hc):
    hf=PatternFill('solid',fgColor=hc);hft=Font(bold=True,color=WHITE,size=11)
    bdr=Border(left=Side(style='thin',color='CCCCCC'),right=Side(style='thin',color='CCCCCC'),
               top=Side(style='thin',color='CCCCCC'),bottom=Side(style='thin',color='CCCCCC'))
    for cell in ws[1]: cell.fill=hf;cell.font=hft;cell.alignment=Alignment(horizontal='center');cell.border=bdr
    ws.row_dimensions[1].height=22
    for i,row in enumerate(ws.iter_rows(min_row=2),start=2):
        rf=PatternFill('solid',fgColor=LGRAY if i%2==0 else WHITE)
        for cell in row: cell.fill=rf;cell.border=bdr;cell.alignment=Alignment(horizontal='center')
    for col in ws.columns:
        mx=max(len(str(c.value or '')) for c in col)+4
        ws.column_dimensions[get_column_letter(col[0].column)].width=min(mx,28)
for sname,hcolor,title in [
    ('KPI Summary','0A0F1E','🎓 Education Analytics — Key Performance Indicators'),
    ('Dept Performance','1A3A5C','📊 Academic Performance by Department'),
    ('Location Analysis','1A3A5C','🌏 Dropout Risk by Student Location'),
    ('Parental Edu Impact','1A3A5C','👨‍👩‍🎓 Impact of Parental Education on Grades'),
]:
    ws=wb[sname]; style_sheet(ws,hcolor)
    ws.insert_rows(1); ws.merge_cells(f'A1:{get_column_letter(ws.max_column)}1')
    tc=ws['A1']; tc.value=title
    tc.font=Font(bold=True,size=14,color=WHITE)
    tc.fill=PatternFill('solid',fgColor='0A0F1E')
    tc.alignment=Alignment(horizontal='center',vertical='center')
    ws.row_dimensions[1].height=30
wb.save(path)
print("✅ STEP 3 COMPLETE — Education_Report.xlsx saved")
